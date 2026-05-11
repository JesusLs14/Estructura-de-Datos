# Estructura-de-Datos/ordenamiento_externo/manejador_archivos.py

import csv
import json
import os
import re
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox

def ventana_seleccion_hojas(hojas: list) -> list:
    """Abre un popup de Tkinter para elegir qué hojas de Excel leer."""
    raiz = tk.Tk()
    raiz.title("Parámetros: Hojas de Excel")
    raiz.attributes("-topmost", True)
    
    tk.Label(raiz, text="El archivo Excel tiene múltiples hojas.\nSelecciona cuáles deseas procesar:", 
             font=("Helvetica", 10, "bold"), padx=15, pady=10).pack()
    
    marco = tk.Frame(raiz, padx=20)
    marco.pack(anchor="w")
    
    variables_casillas = []
    for hoja in hojas:
        variable = tk.IntVar(value=1)
        variables_casillas.append(variable)
        tk.Checkbutton(marco, text=hoja, variable=variable, font=("Helvetica", 10)).pack(anchor="w")
        
    seleccion = []
    def confirmar():
        for i, v in enumerate(variables_casillas):
            if v.get() == 1:
                seleccion.append(hojas[i])
        if not seleccion:
            messagebox.showwarning("Atención", "Debes seleccionar al menos una hoja.")
            return
        raiz.quit()
        
    tk.Button(raiz, text="✔ Cargar Seleccionadas", command=confirmar, bg="#3fb950", fg="white", 
              font=("Helvetica", 10, "bold"), pady=5, padx=10).pack(pady=10)
    
    raiz.eval('tk::PlaceWindow . center')
    raiz.mainloop()
    raiz.destroy()
    return seleccion


def _extraer_numeros_json(datos) -> list:
    """Extrae estrictamente números de estructuras JSON/YAML anidadas."""
    numeros = []
    if isinstance(datos, list):
        for item in datos:
            if isinstance(item, (int, float)):
                numeros.append(int(item))
            elif isinstance(item, str):
                try: 
                    numeros.append(int(float(item.strip())))
                except ValueError: 
                    pass
            elif isinstance(item, (dict, list)):
                numeros.extend(_extraer_numeros_json(item))
    elif isinstance(datos, dict):
        for valor in datos.values(): 
            numeros.extend(_extraer_numeros_json(valor))
    return numeros


def cargar_archivo(ruta: str) -> list:
    """Lee un archivo y extrae una lista de enteros estrictamente."""
    extension = os.path.splitext(ruta)[1].lower()

    if extension == ".txt":
        with open(ruta, encoding="utf-8") as f:
            texto = f.read()
        numeros = re.findall(r"-?\d+(?:\.\d+)?", texto)
        if not numeros: 
            raise ValueError("No se encontraron números en el archivo .txt")
        return [int(float(n)) for n in numeros]

    elif extension == ".csv":
        resultados = []
        with open(ruta, encoding="utf-8") as f:
            lector = csv.reader(f)
            for fila in lector:
                for celda in fila:
                    try: 
                        resultados.append(int(float(celda.strip())))
                    except ValueError: 
                        pass  # Ignora encabezados o texto
        if not resultados: 
            raise ValueError("No se encontraron valores numéricos en el .csv")
        return resultados

    elif extension == ".json":
        with open(ruta, encoding="utf-8") as f:
            datos_json = json.load(f)
        numeros = _extraer_numeros_json(datos_json)
        if not numeros: 
            raise ValueError("No se encontraron números en el .json")
        return numeros

    elif extension == ".xml":
        import xml.etree.ElementTree as ET
        arbol = ET.parse(ruta)
        raiz = arbol.getroot()
        numeros = []
        for etiqueta in ("value", "item", "number", "data", "val", "num"):
            for elemento in raiz.iter(etiqueta):
                try: 
                    numeros.append(int(float(elemento.text.strip())))
                except (TypeError, ValueError, AttributeError): 
                    pass
        if not numeros:
            texto = ET.tostring(raiz, encoding="unicode")
            numeros = [int(float(n)) for n in re.findall(r"-?\d+(?:\.\d+)?", texto)]
        if not numeros: 
            raise ValueError("No se encontraron números en el .xml")
        return numeros

    elif extension in (".yaml", ".yml"):
        try: 
            import yaml
        except ImportError: 
            raise ImportError("Falta PyYAML. Instálalo con: pip install pyyaml")
        with open(ruta, encoding="utf-8") as f:
            datos_yaml = yaml.safe_load(f)
        numeros = _extraer_numeros_json(datos_yaml)
        if not numeros: 
            raise ValueError("No se encontraron números en el .yaml")
        return numeros

    elif extension == ".xlsx":
        try: 
            import openpyxl
        except ImportError: 
            raise ImportError("Falta openpyxl. Instálalo con: pip install openpyxl")
        libro_excel = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        hojas = libro_excel.sheetnames
        seleccionadas = ventana_seleccion_hojas(hojas) if len(hojas) > 1 else hojas
            
        numeros = []
        for hoja in seleccionadas:
            pagina = libro_excel[hoja]
            for fila in pagina.iter_rows(values_only=True):
                for celda in fila:
                    if isinstance(celda, (int, float)):
                        numeros.append(int(celda))
        libro_excel.close()
        if not numeros: 
            raise ValueError("No se encontraron números en las hojas seleccionadas (.xlsx)")
        return numeros

    elif extension == ".xls":
        try: 
            import xlrd
        except ImportError: 
            raise ImportError("Falta xlrd. Instálalo con: pip install xlrd")
        libro_excel = xlrd.open_workbook(ruta)
        hojas = libro_excel.sheet_names()
        seleccionadas = ventana_seleccion_hojas(hojas) if len(hojas) > 1 else hojas
            
        numeros = []
        for hoja in seleccionadas:
            pagina = libro_excel.sheet_by_name(hoja)
            for r in range(pagina.nrows):
                for c in range(pagina.ncols):
                    celda = pagina.cell(r, c)
                    if celda.ctype in (2, 3):
                        numeros.append(int(celda.value))
        if not numeros: 
            raise ValueError("No se encontraron números en las hojas seleccionadas (.xls)")
        return numeros

    elif extension == ".npy":
        arreglo_numpy = np.load(ruta, allow_pickle=False).flatten()
        return [int(x) for x in arreglo_numpy]
    else:
        raise ValueError(f"Extensión '{extension}' no soportada por el sistema.")


def abrir_dialogo_archivo() -> tuple:
    """Abre el explorador de archivos nativo del sistema operativo."""
    raiz = tk.Tk()
    raiz.withdraw()
    raiz.attributes("-topmost", True)

    tipos_soportados = [
        ("Formatos compatibles", "*.txt *.csv *.json *.xml *.yaml *.yml *.xlsx *.xls *.npy"),
        ("Todos los archivos", "*.*"),
    ]
    rutas_seleccionadas = filedialog.askopenfilenames(title="Seleccionar archivo(s) de datos numéricos", filetypes=tipos_soportados)
    raiz.destroy()

    if not rutas_seleccionadas: 
        return None, ""

    datos_totales = []
    nombres_archivos = []
    try:
        for ruta in rutas_seleccionadas:
            datos_extraidos = cargar_archivo(ruta)
            datos_totales.extend(datos_extraidos)
            nombres_archivos.append(os.path.basename(ruta))
            
        return datos_totales, ", ".join(nombres_archivos)
    except Exception as e:
        raiz_error = tk.Tk()
        raiz_error.withdraw()
        messagebox.showerror("Error al procesar", str(e))
        raiz_error.destroy()
        return None, ""